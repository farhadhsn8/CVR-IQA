import torch
import torch.utils.data as data
import torchvision
from PIL import Image
import os
import os.path
import scipy.io
import numpy as np
import csv
import random
from openpyxl import load_workbook
import pandas as pd




def read_csv_column_to_list(file_path, column_name):
    df = pd.read_csv(file_path)
    column_data = df[column_name].tolist()
    return column_data




class Kadid10kFolder(data.Dataset):
    def __init__(self, root, HQ_diff_content_root, index, transform, HQ_diff_content_transform, patch_num, patch_size=224, self_patch_num=10, repeatable_loss = False):
        self.repeatable_loss = repeatable_loss
        self.patch_size = patch_size
        self.self_patch_num = self_patch_num
        self.HQ_diff_content_root = HQ_diff_content_root

        LQ_paths = []
        HQ_paths = []
        mos_all = []
        csv_file = os.path.join(root, 'dmos.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                LQ_paths.append(row['dist_img'])
                HQ_paths.append(row['ref_img'])
                mos = np.array(float(row['dmos'])).astype(np.float32)
                mos_all.append(mos)

        sample = []
        for _, item in enumerate(index):
            for _ in range(patch_num):
                sample.append((os.path.join(root, 'images', LQ_paths[item]),os.path.join(root, 'images', HQ_paths[item]), mos_all[item]))

        self.HQ_diff_content_paths = []
        for HQ_diff_content_img_path in os.listdir(HQ_diff_content_root):
            if HQ_diff_content_img_path[-3:] == 'png' or HQ_diff_content_img_path[-3:] == 'jpg' or HQ_diff_content_img_path[-3:] == 'bmp':
                self.HQ_diff_content_paths.append(os.path.join(HQ_diff_content_root, HQ_diff_content_img_path))

        self.samples = sample
        self.transform = transform
        self.HQ_diff_content_transform = HQ_diff_content_transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, HQ, HQ_diff_content, target) where target is IQA values of the target LQ.
        """
        LQ_path, HQ_path, target = self.samples[index]
        HQ_diff_content_path1 = self.HQ_diff_content_paths[random.randint(0, len(self.HQ_diff_content_paths)-1)]
        HQ_diff_content_path2 = self.HQ_diff_content_paths[random.randint(0, len(self.HQ_diff_content_paths)-1)]
        LQ = pil_loader(LQ_path)
        HQ_diff_content1 = pil_loader(HQ_diff_content_path1)
        HQ_diff_content2 = pil_loader(HQ_diff_content_path2)
        HQ = pil_loader(HQ_path)
        LQ_patches1, HQ_patches, HQ_diff_content_patches1, LQ_patches2, HQ_diff_content_patches2 = [], [], [], [], []
        for _ in range(self.self_patch_num):
            LQ_patch1, HQ_patch = getPairRandomPatch(LQ,HQ, crop_size=self.patch_size)
            LQ_patch2, _ = getPairRandomPatch(LQ,HQ, crop_size=self.patch_size)
            LQ_patch1 = self.transform(LQ_patch1)
            LQ_patch2 = self.transform(LQ_patch2)
            HQ_patch = self.transform(HQ_patch)
            HQ_diff_content_patch1 = self.HQ_diff_content_transform(HQ_diff_content1)
            HQ_diff_content_patch2 = self.HQ_diff_content_transform(HQ_diff_content2)
            LQ_patches1.append(LQ_patch1.unsqueeze(0))
            LQ_patches2.append(LQ_patch2.unsqueeze(0))
            HQ_patches.append(HQ_patch.unsqueeze(0))
            HQ_diff_content_patches1.append(HQ_diff_content_patch1.unsqueeze(0))
            HQ_diff_content_patches2.append(HQ_diff_content_patch2.unsqueeze(0))
        #[self_patch_num, 3, patch_size, patch_size]
        LQ_patches1 = torch.cat(LQ_patches1, 0)
        LQ_patches2 = torch.cat(LQ_patches2, 0)
        HQ_patches = torch.cat(HQ_patches, 0)
        HQ_diff_content_patches1 = torch.cat(HQ_diff_content_patches1, 0)
        HQ_diff_content_patches2 = torch.cat(HQ_diff_content_patches2, 0)

        if self.repeatable_loss == False:
            return LQ_patches1, HQ_patches, HQ_diff_content_patches1, target
        if self.repeatable_loss == True:
            return LQ_patches1, HQ_patches, LQ_patches2, HQ_diff_content_patches1, HQ_diff_content_patches2, target

    def __len__(self):
        length = len(self.samples)
        return length

class LIVEFolder(data.Dataset):

    def __init__(self, root, HQ_diff_content_root, index, transform, HQ_diff_content_transform, patch_num, patch_size=224, self_patch_num=10):
        self.patch_size =patch_size
        self.self_patch_num = self_patch_num
        self.root = root
        self.HQ_diff_content_root = HQ_diff_content_root

        refpath = os.path.join(root, 'refimgs')
        refname = getFileName(refpath, '.bmp')

        jp2kroot = os.path.join(root, 'jp2k')
        jp2kname = self.getDistortionTypeFileName(jp2kroot, 227, islive=True)
        jpegroot = os.path.join(root, 'jpeg')
        jpegname = self.getDistortionTypeFileName(jpegroot, 233,islive=True)
        wnroot = os.path.join(root, 'wn')
        wnname = self.getDistortionTypeFileName(wnroot, 174,islive=True)

        gblurroot = os.path.join(root, 'gblur')
        gblurname = self.getDistortionTypeFileName(gblurroot, 174,islive=True)

        fastfadingroot = os.path.join(root, 'fastfading')
        fastfadingname = self.getDistortionTypeFileName(fastfadingroot, 174,islive=True)

        imgpath = jp2kname + jpegname + wnname + gblurname + fastfadingname

        dmos = scipy.io.loadmat(os.path.join(root, 'dmos_realigned.mat'))
        labels = dmos['dmos_new'].astype(np.float32)

        orgs = dmos['orgs']
        refpaths_all = scipy.io.loadmat(os.path.join(root, 'refnames_all.mat'))
        refpaths_all = refpaths_all['refnames_all']

        sample = []
        for i in range(0, len(index)):
            train_sel = (refname[index[i]] == refpaths_all)
            train_sel = train_sel * ~orgs.astype(np.bool_)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[1].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    LQ_path = imgpath[item]
                    HQ_path = os.path.join(root, 'refimgs', refpaths_all[0][item][0])
                    label = labels[0][item]
                    sample.append((LQ_path, HQ_path, label))
                # print(self.imgpath[item])
        
        self.HQ_diff_content_path = []
        for HQ_diff_content_img_name in os.listdir(HQ_diff_content_root):
            if HQ_diff_content_img_name[-3:] == 'png' or HQ_diff_content_img_name[-3:] == 'jpg' or HQ_diff_content_img_name[-3:] == 'bmp':
                self.HQ_diff_content_path.append(os.path.join(HQ_diff_content_root, HQ_diff_content_img_name))

        self.samples = sample
        self.transform = transform
        self.HQ_diff_content_transform = HQ_diff_content_transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, HQ, HQ_diff_content, target) where target is IQA values of the target LQ.
        """
        LQ_path, HQ_path, target = self.samples[index]
        HQ_diff_content_path = self.HQ_diff_content_path[random.randint(0, len(self.HQ_diff_content_path)-1)]
        LQ = pil_loader(LQ_path)
        HQ = pil_loader(HQ_path)
        HQ_diff_content = pil_loader(HQ_diff_content_path)
        LQ_patches, HQ_patches, HQ_diff_content_patches = [], [], []
        for _ in range(self.self_patch_num):
            LQ_patch, HQ_patch = getPairRandomPatch(LQ, HQ, crop_size=self.patch_size)
            
            LQ_patch = self.transform(LQ_patch)
            HQ_patch = self.transform(HQ_patch)
            HQ_diff_content_patch = self.HQ_diff_content_transform(HQ_diff_content)
            
            LQ_patches.append(LQ_patch.unsqueeze(0))
            HQ_patches.append(HQ_patch.unsqueeze(0))
            HQ_diff_content_patches.append(HQ_diff_content_patch.unsqueeze(0))
        #[self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)
        HQ_patches = torch.cat(HQ_patches, 0)
        HQ_diff_content_patches = torch.cat(HQ_diff_content_patches, 0)

        return LQ_patches, HQ_patches, HQ_diff_content_patches, target

    def __len__(self):
        length = len(self.samples)
        return length

    def getDistortionTypeFileName(self, path, num,islive=False):
        filename = []
        index = 1
        for i in range(0, num):
            if islive:
                name = "img"+str(index)+".bmp"
            else:
                name = '{:0>3d}{}'.format(index, '.bmp')
            filename.append(os.path.join(path, name))
            index = index + 1
        return filename

class CSIQFolder(data.Dataset):

    def __init__(self, root, HQ_diff_content_root, index, transform, HQ_diff_content_transform, patch_num, patch_size =224, self_patch_num=10):
        self.patch_size =patch_size
        self.self_patch_num = self_patch_num

        refpath = os.path.join(root, 'src_imgs')
        refname = getFileName(refpath,'.png')
        txtpath = os.path.join(root, 'scores.txt')
        fh = open(txtpath, 'r')
        LQ_pathes = []
        target = []
        refpaths_all = []
        for line in fh:
            words = line.split(" ")
            LQ_pathes.append((words[3])[9:])
            target.append(words[4][:-2])
            refpaths_all.append(words[2].split("/")[1])

        labels = np.array(target).astype(np.float32)
        refpaths_all = np.array(refpaths_all)
        sample = []
        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refpaths_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    LQ_path = os.path.join(root, 'dst_imgs', LQ_pathes[item])
                    HQ_path = os.path.join(root, 'src_imgs', refpaths_all[item])
                    label = labels[item]
                    sample.append((LQ_path, HQ_path, label))
        
        self.HQ_diff_content = []
        for HQ_diff_content_img_name in os.listdir(HQ_diff_content_root):
            if HQ_diff_content_img_name[-3:] == 'png' or HQ_diff_content_img_name[-3:] == 'jpg' or HQ_diff_content_img_name[-3:] == 'bmp':
                self.HQ_diff_content.append(os.path.join(HQ_diff_content_root, HQ_diff_content_img_name))

        self.samples = sample
        self.transform = transform
        self.HQ_diff_content_transform = HQ_diff_content_transform
    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, HQ, HQ_diff_content, target) where target is IQA values of the target LQ.
        """
        LQ_path, HQ_path, target = self.samples[index]
        HQ_diff_content_path = self.HQ_diff_content[random.randint(0, len(self.HQ_diff_content)-1)]
        LQ = pil_loader(LQ_path)
        HQ = pil_loader(HQ_path)
        HQ_diff_content = pil_loader(HQ_diff_content_path)
        LQ_patches, HQ_patches, HQ_diff_content_patches = [], [], []
        for _ in range(self.self_patch_num):
            LQ_patch, HQ_patch = getPairRandomPatch(LQ, HQ, crop_size=self.patch_size)
            
            LQ_patch = self.transform(LQ_patch)
            HQ_patch = self.transform(HQ_patch)
            HQ_diff_content_patch = self.HQ_diff_content_transform(HQ_diff_content)
            
            LQ_patches.append(LQ_patch.unsqueeze(0))
            HQ_patches.append(HQ_patch.unsqueeze(0))
            HQ_diff_content_patches.append(HQ_diff_content_patch.unsqueeze(0))
        #[self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)
        HQ_patches = torch.cat(HQ_patches, 0)
        HQ_diff_content_patches = torch.cat(HQ_diff_content_patches, 0)

        return LQ_patches, HQ_patches, HQ_diff_content_patches, target

    def __len__(self):
        length = len(self.samples)
        return length

class PIQ23Folder(data.Dataset): # mode: all, train80, test20

    def __init__(self, root, HQ_diff_content_root, index, transform, HQ_diff_content_transform, patch_num, patch_size =224, self_patch_num=10 , mode = "all" , type="Overall"):
        self.patch_size =patch_size
        self.self_patch_num = self_patch_num
        if mode == "all":
            txtpath = os.path.join(root, 'Scores_'+type+'.csv')
        if mode =="test20":
            txtpath = os.path.join(root, 'ntire24_overall_scene_test.csv')
        if mode =="train80":    
            txtpath = os.path.join(root, 'ntire24_overall_scene_train.csv')

        print("PIQ23  |  type" , type)
        print("mode" , mode)
        print("csv file address" , txtpath)

        LQ_pathes = read_csv_column_to_list(column_name="IMAGE PATH" , file_path=txtpath)
        target = read_csv_column_to_list(column_name="JOD" , file_path=txtpath)
 

        labels = np.array(target).astype(np.float32)
        sample = []
        for _, item in enumerate(index):
            for _ in range(patch_num):
                LQ_path = os.path.join(root, LQ_pathes[item])
                LQ_path = LQ_path.replace("\\", "/")
                label = labels[item]
                sample.append((LQ_path, label))
        
        self.HQ_diff_content = []
        for HQ_diff_content_img_name in os.listdir(HQ_diff_content_root):
            if HQ_diff_content_img_name[-3:] == 'png' or HQ_diff_content_img_name[-3:] == 'jpg' or HQ_diff_content_img_name[-3:] == 'bmp':
                self.HQ_diff_content.append(os.path.join(HQ_diff_content_root, HQ_diff_content_img_name))

        self.samples = sample
        self.transform = transform
        self.HQ_diff_content_transform = HQ_diff_content_transform
    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, HQ, HQ_diff_content, target) where target is IQA values of the target LQ.
        """
        LQ_path, target = self.samples[index]
        HQ_diff_content_path = self.HQ_diff_content[random.randint(0, len(self.HQ_diff_content)-1)]
        LQ = pil_loader(LQ_path)
        HQ_diff_content = pil_loader(HQ_diff_content_path)
        LQ_patches,  HQ_diff_content_patches = [], []
        for _ in range(self.self_patch_num):
            LQ_patch = self.HQ_diff_content_transform(LQ)
            HQ_diff_content_patch = self.HQ_diff_content_transform(HQ_diff_content)
            
            LQ_patches.append(LQ_patch.unsqueeze(0))
            HQ_diff_content_patches.append(HQ_diff_content_patch.unsqueeze(0))
        #[self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)
        HQ_diff_content_patches = torch.cat(HQ_diff_content_patches, 0)

        return LQ_patches, _, HQ_diff_content_patches, target

    def __len__(self):
        length = len(self.samples)
        return length


class TID2013Folder(data.Dataset):

    def __init__(self, root, HQ_diff_content_root, index, transform, HQ_diff_content_transform, patch_num, patch_size =224, self_patch_num=10):
        self.patch_size =patch_size
        self.self_patch_num = self_patch_num
        
        refpath = os.path.join(root, 'reference_images')
        refname = self._getTIDFileName(refpath,'.bmp.BMP')
        txtpath = os.path.join(root, 'mos_with_names.txt')
        fh = open(txtpath, 'r')
        LQ_pathes = []
        target = []
        refpaths_all = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            LQ_pathes.append((words[1]))
            target.append(words[0])
            ref_temp = words[1].split("_")
            refpaths_all.append(ref_temp[0][1:])
        labels = np.array(target).astype(np.float32)
        refpaths_all = np.array(refpaths_all)

        sample = []
        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refpaths_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    LQ_path = os.path.join(root, 'distorted_images', LQ_pathes[item])
                    refHQ_name = 'I' + LQ_pathes[item].split("_")[0][1:] + '.BMP'
                    HQ_path = os.path.join(refpath, refHQ_name)
                    label = labels[item]
                    sample.append((LQ_path, HQ_path, label))
        self.HQ_diff_content = []
        for HQ_diff_content_img_name in os.listdir(HQ_diff_content_root):
            if HQ_diff_content_img_name[-3:] == 'png' or HQ_diff_content_img_name[-3:] == 'jpg' or HQ_diff_content_img_name[-3:] == 'bmp':
                self.HQ_diff_content.append(os.path.join(HQ_diff_content_root, HQ_diff_content_img_name))
       
        self.samples = sample
        self.transform = transform
        self.HQ_diff_content_transform = HQ_diff_content_transform

    def _getTIDFileName(self, path, suffix):
        filename = []
        f_list = os.listdir(path)
        for i in f_list:
            if suffix.find(os.path.splitext(i)[1]) != -1:
                filename.append(i[1:3])
        return filename

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, HQ, HQ_diff_content, target) where target is IQA values of the target LQ.
        """
        LQ_path, HQ_path, target = self.samples[index]
        HQ_diff_content_path = self.HQ_diff_content[random.randint(0, len(self.HQ_diff_content)-1)]
        LQ = pil_loader(LQ_path)
        HQ = pil_loader(HQ_path)
        HQ_diff_content = pil_loader(HQ_diff_content_path)
        LQ_patches, HQ_patches, HQ_diff_content_patches = [], [], []
        for _ in range(self.self_patch_num):
            LQ_patch, HQ_patch = getPairRandomPatch(LQ, HQ, crop_size=self.patch_size)
            
            LQ_patch = self.transform(LQ_patch)
            HQ_patch = self.transform(HQ_patch)
            HQ_diff_content_patch = self.HQ_diff_content_transform(HQ_diff_content)
            
            LQ_patches.append(LQ_patch.unsqueeze(0))
            HQ_patches.append(HQ_patch.unsqueeze(0))
            HQ_diff_content_patches.append(HQ_diff_content_patch.unsqueeze(0))
        #[self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)
        HQ_patches = torch.cat(HQ_patches, 0)
        HQ_diff_content_patches = torch.cat(HQ_diff_content_patches, 0)

        return LQ_patches, HQ_patches, HQ_diff_content_patches, target

    def __len__(self):
        length = len(self.samples)
        return length

class LIVEChallengeFolder(data.Dataset):
    def __init__(self, root, HQ_diff_content_root, index, transform, HQ_diff_content_transform, patch_num, patch_size =224, self_patch_num=10):
        self.patch_size =patch_size
        self.self_patch_num = self_patch_num

        LQ_pathes = scipy.io.loadmat(os.path.join(root, 'Data', 'AllImages_release.mat'))
        LQ_pathes = LQ_pathes['AllImages_release']
        LQ_pathes = LQ_pathes[7:1169]
        mos = scipy.io.loadmat(os.path.join(root, 'Data', 'AllMOS_release.mat'))
        labels = mos['AllMOS_release'].astype(np.float32)
        labels = labels[0][7:1169]

        sample = []
        for _, item in enumerate(index):
            for _ in range(patch_num):
                sample.append((os.path.join(root, 'Images', LQ_pathes[item][0][0]), labels[item]))
        
        self.HQ_diff_content_paths = []
        for HQ_diff_content_img_path in os.listdir(HQ_diff_content_root):
            if HQ_diff_content_img_path[-3:] == 'png' or HQ_diff_content_img_path[-3:] == 'jpg' or HQ_diff_content_img_path[-3:] == 'bmp':
                self.HQ_diff_content_paths.append(os.path.join(HQ_diff_content_root, HQ_diff_content_img_path))

        self.samples = sample
        self.transform = transform
        self.HQ_diff_content_transform = HQ_diff_content_transform
    
    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, _, HQ_diff_content, target) where target is IQA values of the target LQ.
        """
        LQ_path, target = self.samples[index]
        HQ_diff_content_path = self.HQ_diff_content_paths[random.randint(0, len(self.HQ_diff_content_paths)-1)]
        LQ = pil_loader(LQ_path)
        HQ_diff_content = pil_loader(HQ_diff_content_path)
        LQ_patches, HQ_diff_content_patches = [], []
        for _ in range(self.self_patch_num):
            LQ_patch = self.HQ_diff_content_transform(LQ)
            HQ_diff_content_patch = self.HQ_diff_content_transform(HQ_diff_content)
            
            LQ_patches.append(LQ_patch.unsqueeze(0))
            HQ_diff_content_patches.append(HQ_diff_content_patch.unsqueeze(0))
        #[self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)
        HQ_diff_content_patches = torch.cat(HQ_diff_content_patches, 0)

        return LQ_patches, _, HQ_diff_content_patches, target
    
    def __len__(self):
        length = len(self.samples)
        return length

class BIDChallengeFolder(data.Dataset):
    def __init__(self, root, HQ_diff_content_root, index, transform, HQ_diff_content_transform, patch_num, patch_size =224, self_patch_num=10):
        self.patch_size =patch_size
        self.self_patch_num = self_patch_num

        LQ_pathes = []
        labels = []

        xls_file = os.path.join(root, 'DatabaseGrades.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        count = 1
        for _ in rows:
            count += 1
            img_num = (booksheet.cell(row=count, column=1).value)
            img_name = "DatabaseImage%04d.JPG" % (img_num)
            LQ_pathes.append(img_name)
            mos = (booksheet.cell(row=count, column=2).value)
            mos = np.array(mos)
            mos = mos.astype(np.float32)
            labels.append(mos)
            if count == 587:
                break

        sample = []
        for _, item in enumerate(index):
            for _ in range(patch_num):
                sample.append((os.path.join(root, LQ_pathes[item]), labels[item]))
        
        self.HQ_diff_content_paths = []
        for HQ_diff_content_img_path in os.listdir(HQ_diff_content_root):
            if HQ_diff_content_img_path[-3:] == 'png' or HQ_diff_content_img_path[-3:] == 'jpg' or HQ_diff_content_img_path[-3:] == 'bmp':
                self.HQ_diff_content_paths.append(os.path.join(HQ_diff_content_root, HQ_diff_content_img_path))

        self.samples = sample
        self.transform = transform
        self.HQ_diff_content_transform = HQ_diff_content_transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, _, HQ_diff_content, target) where target is IQA values of the target LQ.
        """
        LQ_path, target = self.samples[index]
        HQ_diff_content_path = self.HQ_diff_content_paths[random.randint(0, len(self.HQ_diff_content_paths)-1)]
        LQ = pil_loader(LQ_path)
        HQ_diff_content = pil_loader(HQ_diff_content_path)
        LQ_patches, HQ_diff_content_patches = [], []
        for _ in range(self.self_patch_num):
            LQ_patch = self.HQ_diff_content_transform(LQ)
            HQ_diff_content_patch = self.HQ_diff_content_transform(HQ_diff_content)
            
            LQ_patches.append(LQ_patch.unsqueeze(0))
            HQ_diff_content_patches.append(HQ_diff_content_patch.unsqueeze(0))
        #[self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)
        HQ_diff_content_patches = torch.cat(HQ_diff_content_patches, 0)

        return LQ_patches, _, HQ_diff_content_patches, target
    
    def __len__(self):
        length = len(self.samples)
        return length

class Koniq_10kFolder(data.Dataset):
    def __init__(self, root, HQ_diff_content_root, index, transform, HQ_diff_content_transform, patch_num, patch_size =224, self_patch_num=10):
        self.patch_size =patch_size
        self.self_patch_num = self_patch_num

        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'koniq10k_scores_and_distributions.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['image_name'])
                mos = np.array(float(row['MOS_zscore'])).astype(np.float32)
                mos_all.append(mos)

        sample = []
        for _, item in enumerate(index):
            for _ in range(patch_num):
                sample.append((os.path.join(root, '1024x768', imgname[item]), mos_all[item]))
        
        self.HQ_diff_content_paths = []
        for HQ_diff_content_img_path in os.listdir(HQ_diff_content_root):
            if HQ_diff_content_img_path[-3:] == 'png' or HQ_diff_content_img_path[-3:] == 'jpg' or HQ_diff_content_img_path[-3:] == 'bmp':
                self.HQ_diff_content_paths.append(os.path.join(HQ_diff_content_root, HQ_diff_content_img_path))


        self.samples = sample
        self.transform = transform
        self.HQ_diff_content_transform = HQ_diff_content_transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, _, HQ_diff_content, target) where target is IQA values of the target LQ.
        """
        LQ_path, target = self.samples[index]
        HQ_diff_content_path = self.HQ_diff_content_paths[random.randint(0, len(self.HQ_diff_content_paths)-1)]
        LQ = pil_loader(LQ_path)
        HQ_diff_content = pil_loader(HQ_diff_content_path)
        LQ_patches, HQ_diff_content_patches = [], []
        for _ in range(self.self_patch_num):
            LQ_patch = self.HQ_diff_content_transform(LQ)
            HQ_diff_content_patch = self.HQ_diff_content_transform(HQ_diff_content)
            
            LQ_patches.append(LQ_patch.unsqueeze(0))
            HQ_diff_content_patches.append(HQ_diff_content_patch.unsqueeze(0))
        #[self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)
        HQ_diff_content_patches = torch.cat(HQ_diff_content_patches, 0)

        return LQ_patches, _, HQ_diff_content_patches, target

    def __len__(self):
        length = len(self.samples)
        return length

def getFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if os.path.splitext(i)[1] == suffix:
            filename.append(i)
    return filename


def getPairRandomPatch(img1, img2, crop_size=512):
    (iw,ih) = img1.size
    # print(ih,iw)

    ip = int(crop_size)

    ix = random.randrange(0, iw - ip + 1)
    iy = random.randrange(0, ih - ip + 1)


    img1_patch = img1.crop((ix, iy, ix+ip, iy+ip))#左上右下
    img2_patch = img2.crop((ix, iy, ix+ip, iy+ip))#左上右下

    return img1_patch, img2_patch

def getPairAugment(img1, img2, hflip=True, vflip=False, rot=False):
    hflip = hflip and random.random() < 0.5
    vflip = vflip and random.random() < 0.5
    rot180 = rot and random.random() < 0.5

    if hflip: 
        img1 = img1.transpose(Image.FLIP_TOP_BOTTOM)
        img2 = img2.transpose(Image.FLIP_TOP_BOTTOM)
    if vflip: 
        img1 = img1.transpose(Image.FLIP_LEFT_RIGHT)
        img2 = img2.transpose(Image.FLIP_LEFT_RIGHT)
    if rot180: 
        img1 = img1.transpose(Image.ROTATE_180)
        img2 = img2.transpose(Image.ROTATE_180)
        
    return img1, img2


def getSelfPatch(img, patch_size, patch_num, is_random=True):
    (iw,ih) = img.size
    patches = []
    for i in range(patch_num):
        if is_random:
            ix = random.randrange(0, iw - patch_size + 1)
            iy = random.randrange(0, ih - patch_size + 1)
        else:ix,iy=(iw - patch_size + 1)//2,(ih - patch_size + 1)//2

        # patch = img[iy:iy + lr_size, ix:ix + lr_size, :]#上下左右
        patch = img.crop((ix, iy, ix+patch_size, iy+patch_size))#左上右下
        patches.append(patch)

    return patches


def pil_loader(path):
    with open(path, 'rb') as f:
        img = Image.open(f)
        return img.convert('RGB')
